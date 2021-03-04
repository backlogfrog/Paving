#extract data from a spreadsheet and convert to Paver XML data for import

#finding files by extension in main dir/opening files, exec .py files
import os
import glob

#justprogressbarthings
from alive_progress import alive_bar
#from progress.bar import Bar


#needed to parse date in scratchXml.py
from datetime import datetime

#colored text for funskies while I wait for data
import colorama
from colorama import Fore, Back, Style 
colorama.init()
#reset color/style after printing
colorama.init(autoreset=True)
#Colors, yay...
#Format: print(Fore.BLACK + "TEXT " + str(VARIABLE))
#Fore: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
#Back: BLACK, RED, GREEN, YELLOW, BLUE, MAGENTA, CYAN, WHITE, RESET.
#Style: DIM, NORMAL, BRIGHT, RESET_ALL

#openpyxl package opens excel sheets
from openpyxl import load_workbook

#NEW DATA HAS NO DATE - NEED TO SET DATE

##################################################################################
##################################################################################
##################################################################################
#set start positions to read data: 2012.xlsx data starts at row 4
#RowIncr=4
RowIncr=2   
##################################################################################
##################################################################################
##################################################################################
##################################################################################
##################################################################################




#import class info from inspectionClasses.py for inspection data and from mapping.py based on column to be printed to XML

from mapping import INSPECTED_SIZE, INSPECTED_DATE, INSPECTED_PID1, INSPECTED_PID2, DCOMMENT, P_LENGTH, P_WIDTH, SAMPLENUMBER, SWEATHERING_CODE, SWEATHERING_S, SWEATHERING_Q, ALLIGATOR_CODE, ALLIGATOR_S, ALLIGATOR_Q, BLOCKCRACK_CODE, BLOCKCRACK_S, BLOCKCRACK_Q, TRANSVERSE__CODE, TRANSVERSE_S, TRANSVERSE_Q, DEPRESSION_CODE, DEPRESSION_S, DEPRESSION_Q, POTHOLE_CODE, POTHOLE_S, POTHOLE_Q, EDGECRACKING_CODE, EDGECRACKING_S, EDGECRACKING_Q, JOINTSPALLING_CODE, JOINTSPALLING_S, JOINTSPALLING_Q, DURABILITYCRACKING_CODE, DURABILITYCRACKING_S, DURABILITYCRACKING_Q, FAULTING_CODE, FAULTING_S, FAULTING_Q, PATCHING_CODE, PATCHING_S, PATCHING_Q, BUMPSAG_CODE, BUMPSAG_S, BUMPSAG_Q, SAMPLESIZE

#set header/schema for xml
xml_header = "<?xml version=\"1.0\" encoding=\"utf-8\" ?>"
xml_schema = "<pavementData xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xsi:noNamespaceSchemaLocation=\"PavementInspectionDataV2.xsd\">"


#############################open workbook from xlsx files in main dir for mult choice
#############################

excel_list = [f for f in glob.glob("*.xlsx")]
print(Fore.BLUE + "Available Spreadsheets:")
for i, db_name in enumerate(excel_list, start=1):
    print ('{}. {}'.format(i, db_name))
  
while True:
	try:
		selected = int(input(Fore.CYAN + 'Select spreadsheet (1-{}): '.format(i)))
		db_name = excel_list[selected-1]
		print(Style.DIM + 'Loading {}'.format(db_name), "\n")
		break
	except (ValueError, IndexError):
		print(Fore.BLUE + 'Invalid. Please enter number between 1 and {}!'.format(i))
		print("	")

#load input/workbook from above and set current active sheet to "sheet"
workbook = load_workbook(filename=db_name, data_only=True)
workbook.sheetnames
sheet = workbook.active


#set/show last row and column to avoid out of range error
#LastRow=sheet.max_row
#LastColumn=sheet.max_column
#print ("Last Column ", LastColumn)


#something is fucky here, last row shows is greater than 
LastRow=sheet.max_row-RowIncr
print("Total row number is ", LastRow)




#LastRow = Copy of 2012.xlsx 685 readable lines
#2012.xlsx has 686 readable lines
FinalRow = sheet.max_row

#check to see if last row has data for first column - if 0 then subtract 1 from last Row
#to avoid empty data/date error
#else, add LastRow to RowIncr(sheet actual-data-start offset)

#DETERMINE IF FURTHER ROWS ARE BLANK/0 AND stop pulling data, im not sure how this works currently 2.19.20
cellCheck=sheet.cell(row=FinalRow, column=1).value
if cellCheck == 0:
	LastRow=LastRow+RowIncr-1
else:
	LastRow=LastRow+RowIncr


#cut filename of .xml/.log file to 4 char
fileName=db_name[0:4]

#create folders for xml and log files
if not os.path.exists('XML'):
    os.makedirs('XML')

if not os.path.exists('LOG'):
    os.makedirs('LOG')

#write files/delete if present
if os.path.exists("XML/"+fileName+".xml"):
  os.remove("XML/"+fileName+".xml")
if os.path.exists("LOG/"+fileName+".log"):
  os.remove("LOG/"+fileName+".log")
  print(Style.DIM + Fore.RED +"Original XML/LOG Deleted \n \n")
else:
  print(Style.DIM + Fore.BLUE +"New file created \n \n")
f = open("XML/"+fileName + ".xml", "a+")
logFile = open("LOG/"+fileName + ".log", "a+")






#add headers to top of page
print(xml_header, "\n", xml_schema, sep="", file=f)

#how many rows read/written
rowsRead=0
ticker = 0

def emptyData():
	print ("Empty: ", row[INSPECTED_PID2],":", row[SAMPLENUMBER], " ###################", file=logFile)
	#print ("Empty: ", row[INSPECTED_PID2],":", row[SAMPLENUMBER], " ###################")

def fullData():
		print("Data: ", row[INSPECTED_PID2], ":", row[SAMPLENUMBER], file=logFile)
		#print("Data: ", row[INSPECTED_PID2], ":", row[SAMPLENUMBER])

#check each code to ensure there is a distress for this row, set ticker to 1 to write
def codeCheck(code):
	
	try:
		if float(code) > 0:
			global ticker
			ticker = 1
			fullData()
			
		else:
			emptyData()
			
	except ValueError:
		
		
		emptyData()
		ticker = 0

#TESTING for specific errors in specific PID import
#PIDREQUEST = input("Enter your PID for PID specific xml: ") 
#print("PID: ", PIDREQUEST) 
#for if statement below: if ticker == 1 and row[INSPECTED_PID2] == int(PIDREQUEST)




        



with alive_bar(LastRow, bar = 'bubbles', spinner = 'notes2') as bar:
	for row in sheet.iter_rows(min_row=RowIncr, max_row=LastRow, values_only=True):
			rowsRead=rowsRead+1
			ticker = 0
			codeCheck(row[SWEATHERING_CODE])	 	
			codeCheck(row[ALLIGATOR_CODE])
			codeCheck(row[BLOCKCRACK_CODE])
			codeCheck(row[TRANSVERSE__CODE])
			codeCheck(row[DEPRESSION_CODE])
			codeCheck(row[POTHOLE_CODE])
			codeCheck(row[EDGECRACKING_CODE])
			codeCheck(row[JOINTSPALLING_CODE])
			codeCheck(row[DURABILITYCRACKING_CODE])
			codeCheck(row[FAULTING_CODE])
			codeCheck(row[PATCHING_CODE])
			codeCheck(row[BUMPSAG_CODE])
			if ticker == 1:
				exec(open("scratchXml.py").read())
			bar()
			#bar.next()
			



#print (Fore.RED + "FIRST SS IS IN 2010 - NO SET DATE \n \n")
	
#close xml main tag and workbook
print("</pavementData>", sep="", file=f)
f.close()
logFile.close()

#print rows and files read
print(Fore.MAGENTA + "\nRows Read: ", int(rowsRead)-1)		
print(Fore.MAGENTA + "File Opened:", db_name)
print(Fore.MAGENTA + "Files Written:", fileName+".xml"+", "+fileName+".log")